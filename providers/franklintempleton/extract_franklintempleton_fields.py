"""Extract selected ETF fields from a downloaded Franklin Templeton UCITS ETF workbook."""

from __future__ import annotations

import argparse
import csv
import os
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR
OUTPUT_DIR = BASE_DIR
RUN_FOLDER_ENV_VAR = "ETF_PIPELINE_RUN_FOLDER"

ISSUER = "Franklin Templeton"

EQUITY_SHEET = "Equity"
FIXED_INCOME_SHEET = "Fixed Income"
ALL_SHEETS = [EQUITY_SHEET, FIXED_INCOME_SHEET]

DATE_ROW = 3
HEADER_ROW = 5
DATA_START_ROW = 6

SOURCE_COLUMNS = {
    "fund_name": "Fund Name",
    "isin": "ISIN",
    "ter": "(TER) Total Expense Ratio",
    "currency": "Currency Code",
    "aum": "Total Net Assets",
}

OUTPUT_COLUMNS = [
    "ETF Name",
    "Issuer",
    "ISIN",
    "CCY",
    "TER(bps)",
    "AUM(M)",
    "Date",
]

ISIN_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{9}\d$")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract selected ETF fields from a Franklin Templeton UCITS ETF .xlsx file."
    )
    parser.add_argument(
        "--input",
        type=Path,
        help="Path to the downloaded .xlsx file. Defaults to the latest file found under the script directory.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Destination CSV path. Defaults to a dated folder under the script directory.",
    )
    parser.add_argument(
        "--sheet",
        choices=["equity", "fixed_income", "all"],
        default="all",
        help="Which sheet(s) to process. Default: all.",
    )
    return parser.parse_args()


def build_run_output_dir(base_dir: Path) -> Path:
    run_folder_name = os.environ.get(RUN_FOLDER_ENV_VAR)

    if run_folder_name:
        output_dir = base_dir / run_folder_name
        output_dir.mkdir(parents=True, exist_ok=True)
        return output_dir

    run_date = datetime.now().strftime("%Y-%m-%d")
    output_dir = base_dir / run_date

    suffix = 1
    while output_dir.exists():
        output_dir = base_dir / f"{run_date} ({suffix})"
        suffix += 1

    output_dir.mkdir(parents=True, exist_ok=False)
    os.environ[RUN_FOLDER_ENV_VAR] = output_dir.name

    return output_dir


def find_latest_download(input_dir: Path) -> Path:
    candidates = sorted(
        (p for p in input_dir.rglob("*.xlsx") if p.is_file()),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )

    if not candidates:
        raise FileNotFoundError(f"No .xlsx files found under {input_dir}")

    return candidates[0]


def build_output_path(output_dir: Path) -> Path:
    dated_output_dir = build_run_output_dir(output_dir)
    return dated_output_dir / "franklintempleton_selected_fields.csv"


def clean_text(value: object | None) -> str:
    if value is None:
        return ""

    text = str(value).strip()

    if text in {"", "-", "—", "- ", " -"}:
        return ""

    return text


def is_valid_isin(value: object | None) -> bool:
    isin = clean_text(value).upper()
    return bool(ISIN_RE.fullmatch(isin))


def normalize_number_text(value: object | None) -> str:
    cleaned = clean_text(value)

    if not cleaned:
        return ""

    cleaned = cleaned.replace("%", "").strip()

    numeric = re.sub(r"[^\d.,-]", "", cleaned)

    if not numeric:
        return ""

    # Example: 1,234.56 -> 1234.56
    if "," in numeric and "." in numeric:
        numeric = numeric.replace(",", "")

    # Example: 1234,56 -> 1234.56
    elif "," in numeric and "." not in numeric:
        numeric = numeric.replace(",", ".")

    return numeric


def format_decimal(value: object | None, places: int = 2) -> str:
    numeric = normalize_number_text(value)

    if not numeric:
        return ""

    try:
        d = Decimal(numeric)
    except InvalidOperation:
        return clean_text(value)

    quantizer = Decimal("1." + "0" * places)
    return format(d.quantize(quantizer, rounding=ROUND_HALF_UP), f".{places}f")


def format_ter(value: object | None) -> str:
    """
    Convert TER to basis points.

    Franklin workbook usually stores TER as percentage units:
    0.09 means 0.09%, so:
    0.09 * 100 = 9.00 bps
    """
    numeric = normalize_number_text(value)

    if not numeric:
        return ""

    try:
        bps = Decimal(numeric) * Decimal("100")
    except InvalidOperation:
        return clean_text(value)

    return format_decimal(str(bps), places=2)


def format_aum(value: object | None) -> str:
    """
    Convert AUM like '$131.79 Million' or '€216.80 Million'
    into numeric millions.

    Example:
    '$131.79 Million' -> '131.79'
    """
    cleaned = clean_text(value)

    if not cleaned:
        return ""

    cleaned = (
        cleaned.replace("Million", "")
        .replace("million", "")
        .replace("MILLION", "")
    )

    return format_decimal(cleaned, places=2)


def format_date(value: object | None) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")

    cleaned = clean_text(value)

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(cleaned, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue

    return cleaned


def extract_file_date(input_path: Path) -> str:
    """
    Read the 'as of <date>' value from row 3.
    Falls back to the parent folder name or today's date.
    """
    try:
        wb = load_workbook(input_path, read_only=True, data_only=True)
        ws = wb[EQUITY_SHEET] if EQUITY_SHEET in wb.sheetnames else wb.active

        for row in ws.iter_rows(min_row=DATE_ROW, max_row=DATE_ROW, values_only=True):
            for cell in row[1:]:
                if cell is not None:
                    result = format_date(cell)
                    if result:
                        wb.close()
                        return result

        wb.close()

    except Exception:
        pass

    parent_match = re.match(r"(\d{4}-\d{2}-\d{2})", input_path.parent.name)

    if parent_match:
        try:
            return datetime.strptime(parent_match.group(1), "%Y-%m-%d").strftime("%d/%m/%Y")
        except ValueError:
            pass

    return datetime.now().strftime("%d/%m/%Y")


def read_sheet(
    input_path: Path,
    sheet_name: str,
    file_date: str,
) -> list[dict[str, str]]:
    """
    Read one workbook sheet and return only real ETF rows.
    Footnotes/legal rows are skipped because they do not have valid ISINs.
    """
    wb = load_workbook(input_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        available_sheets = wb.sheetnames
        wb.close()
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {input_path}. "
            f"Available sheets: {available_sheets}"
        )

    ws = wb[sheet_name]

    header_map: dict[str, int] = {}

    for row in ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True):
        for col_idx, cell_value in enumerate(row):
            if cell_value is not None:
                header_map[str(cell_value).strip()] = col_idx
        break

    missing = [column for column in SOURCE_COLUMNS.values() if column not in header_map]

    if missing:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}': missing expected columns: {missing}")

    def get(row_tuple: tuple, source_column: str) -> object | None:
        idx = header_map.get(source_column)

        if idx is None or idx >= len(row_tuple):
            return None

        return row_tuple[idx]

    rows: list[dict[str, str]] = []

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        fund_name = clean_text(get(row, SOURCE_COLUMNS["fund_name"]))
        isin = clean_text(get(row, SOURCE_COLUMNS["isin"])).upper()

        if not fund_name:
            continue

        # This is the important fix:
        # Skip Footnotes / Legal Information / empty footer rows.
        if not is_valid_isin(isin):
            continue

        record = {
            "ETF Name": fund_name,
            "Issuer": ISSUER,
            "ISIN": isin,
            "CCY": clean_text(get(row, SOURCE_COLUMNS["currency"])).upper(),
            "TER(bps)": format_ter(get(row, SOURCE_COLUMNS["ter"])),
            "AUM(M)": format_aum(get(row, SOURCE_COLUMNS["aum"])),
            "Date": file_date,
        }

        rows.append(record)

    wb.close()
    return rows


def write_csv(output_path: Path, rows: list[dict[str, str]]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def extract_rows(
    input_path: Path | None = None,
    *,
    sheets: list[str] | None = None,
) -> list[dict[str, str]]:
    """
    Return extracted rows without writing to disk.
    Useful for pipeline composition.
    """
    resolved_input = input_path.resolve() if input_path else find_latest_download(INPUT_DIR)
    file_date = extract_file_date(resolved_input)
    target_sheets = sheets or ALL_SHEETS

    all_rows: list[dict[str, str]] = []

    for sheet_name in target_sheets:
        all_rows.extend(read_sheet(resolved_input, sheet_name, file_date))

    return all_rows


def parse_snapshot_rows(input_path: Path) -> list[dict[str, str]]:
    return extract_rows(input_path)


def process_file(
    input_path: Path | None = None,
    output_path: Path | None = None,
    *,
    sheets: list[str] | None = None,
) -> Path:
    resolved_input = input_path.resolve() if input_path else find_latest_download(INPUT_DIR)
    resolved_output = output_path.resolve() if output_path else build_output_path(OUTPUT_DIR)

    rows = extract_rows(resolved_input, sheets=sheets)
    write_csv(resolved_output, rows)

    print("=" * 60)
    print("Franklin Templeton Fields Extractor")
    print("=" * 60)
    print(f"Source file : {resolved_input}")
    print(f"Rows written: {len(rows):,}")
    print(f"Output file : {resolved_output}")
    print("=" * 60)

    return resolved_output


def main() -> None:
    args = parse_args()

    sheet_map = {
        "equity": [EQUITY_SHEET],
        "fixed_income": [FIXED_INCOME_SHEET],
        "all": ALL_SHEETS,
    }

    selected_sheets = sheet_map[args.sheet]

    process_file(args.input, args.output, sheets=selected_sheets)


if __name__ == "__main__":
    main()
