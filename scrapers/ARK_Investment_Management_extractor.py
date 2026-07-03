# scrapers/ark_extractor.py
"""
ARK Invest Europe ETF scraper
Target : https://europe.ark-funds.com/funds/
Output : providers/ARK_Investment_Management/{date}/ark_etfs_{date}.xlsx

Table lives at:  table.ps_etf_tables__table  (inside div#overview)
Rows:            tbody tr[role="row"]
Columns (left→right in DOM):
  0  Name + href  →  fund_name, fund_url
  1  Base Code    →  base_code
  2  ISIN         →  isin
  3  SFDR         →  sfdr_classification
  4  TER (%)      →  ter_pct
  5  AUM ($USD)   →  aum_usd  +  ccy parsed from column header text
  6  Factsheet    →  factsheet_url
"""

import argparse
import asyncio
import os
import re
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError


# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

URL               = "https://europe.ark-funds.com/funds/"
PROVIDER_FOLDER   = "ARK_Investment_Management"
RUN_FOLDER_ENV    = "ETF_PIPELINE_RUN_FOLDER"

PAGE_LOAD_TIMEOUT = 45_000
BASE_URL          = "https://europe.ark-funds.com"
DEFAULT_CCY       = "USD"   # fallback if header parse fails
OUTPUT_DIR        = Path(__file__).resolve().parents[1] / "providers" / PROVIDER_FOLDER

COLUMNS = [
    "fund_name",
    "fund_url",
    "base_code",
    "isin",
    "sfdr_classification",
    "ter_pct",
    "aum_usd",
    "ccy",
    "factsheet_url",
    "scrape_date",
]

HEADER_FILL  = PatternFill("solid", fgColor="1F1F5E")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill("solid", fgColor="F0F0F8")


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Scrape ARK Invest Europe ETF listings.")
    p.add_argument("--headed", action="store_true", default=False,
                   help="Show the browser window (default: headless).")
    return p.parse_args()


# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT FOLDER  –  providers/ARK_Investment_Management/{date}/
# ─────────────────────────────────────────────────────────────────────────────

def build_run_output_dir() -> Path:
    """
    Respects ETF_PIPELINE_RUN_FOLDER as a run-folder name.
    Falls back to providers/ARK_Investment_Management/{today}/ and reuses that same folder on reruns.
    """
    run_folder_name = os.environ.get(RUN_FOLDER_ENV)
    if run_folder_name:
        save_dir = OUTPUT_DIR / run_folder_name
    else:
        save_dir = OUTPUT_DIR / date.today().isoformat()
        os.environ[RUN_FOLDER_ENV] = save_dir.name

    save_dir.mkdir(parents=True, exist_ok=True)
    return save_dir


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def find_first_visible_locator(page, selectors: list[str]):
    for sel in selectors:
        loc = page.locator(sel).first
        try:
            if loc.is_visible(timeout=2_000):
                return loc
        except PlaywrightTimeoutError:
            continue
    return None


def click_with_fallback(page, selectors: list[str], label: str) -> bool:
    loc = find_first_visible_locator(page, selectors)
    if loc:
        loc.click()
        print(f"  ✓ Clicked {label}.")
        return True
    print(f"  ⚠ {label} not found — skipping.")
    return False


def resolve_href(href: str | None) -> str:
    if not href:
        return ""
    return href if href.startswith("http") else BASE_URL + href


def parse_ccy_from_header(header_text: str) -> str:
    """
    Extract ISO-4217 code from the AUM <th> text.
    'AUM ($USD)' → 'USD'   |   'AUM (€EUR)' → 'EUR'
    Falls back to DEFAULT_CCY if nothing is found.
    """
    m = re.search(r"\b([A-Z]{3})\b", header_text)
    return m.group(1) if m else DEFAULT_CCY


# ─────────────────────────────────────────────────────────────────────────────
# GATE / COOKIE DISMISSAL
# ─────────────────────────────────────────────────────────────────────────────

def dismiss_cookie_banner(page) -> None:
    print("  [step] Checking for cookie banner…")
    result = page.evaluate("""
        () => {
            const candidates = [
                document.querySelector('#onetrust-accept-btn-handler'),
                Array.from(document.querySelectorAll('button')).find(b =>
                    b.innerText.toLowerCase().includes('accept all cookies')),
                Array.from(document.querySelectorAll('button')).find(b =>
                    b.innerText.trim().toLowerCase() === 'accept all'),
                Array.from(document.querySelectorAll('button')).find(b =>
                    b.innerText.trim().toLowerCase().includes('accept')),
            ].filter(Boolean);
            const btn = candidates[0];
            if (!btn) return 'NOT_FOUND';
            ['mousedown','mouseup','click'].forEach(t =>
                btn.dispatchEvent(new MouseEvent(t, {bubbles:true, cancelable:true}))
            );
            return 'CLICKED: ' + btn.innerText.trim().slice(0, 60);
        }
    """)
    print(f"    → {result}")
    if result and result.startswith("CLICKED"):
        page.wait_for_timeout(1_500)


def dismiss_investor_gate(page) -> None:
    print("  [step] Checking for investor / T&C gate…")
    gate_texts = [
        "i am not a us person", "i confirm i am not", "confirm and proceed",
        "confirm and continue", "accept and continue", "i accept",
        "agree and continue", "agree", "proceed", "continue", "enter site",
    ]
    result = page.evaluate("""
        (texts) => {
            const btns = Array.from(document.querySelectorAll('button'))
                .filter(b => b.offsetParent !== null);
            for (const text of texts) {
                const btn = btns.find(b =>
                    b.innerText.trim().toLowerCase().includes(text));
                if (btn) {
                    ['mousedown','mouseup','click'].forEach(t =>
                        btn.dispatchEvent(new MouseEvent(t, {bubbles:true, cancelable:true}))
                    );
                    return 'CLICKED: ' + btn.innerText.trim().slice(0, 60);
                }
            }
            return 'NOT_FOUND';
        }
    """, gate_texts)
    print(f"    → {result}")
    if result and result.startswith("CLICKED"):
        page.wait_for_timeout(1_500)


# ─────────────────────────────────────────────────────────────────────────────
# WAIT FOR TABLE
# ─────────────────────────────────────────────────────────────────────────────

def wait_for_fund_table(page) -> None:
    print("  [step] Waiting for ETF table to render…")
    signals = [
        "table.ps_etf_tables__table tbody tr[role='row']",
        "#overview table tbody tr",
        "table.ps_etf_tables__table",
    ]
    for sig in signals:
        try:
            page.wait_for_selector(sig, timeout=PAGE_LOAD_TIMEOUT)
            print(f"    ✓ Table confirmed via: {sig!r}")
            return
        except PlaywrightTimeoutError:
            continue
    raise RuntimeError(
        "ETF table did not appear within timeout.\n"
        "Run with --headed to inspect the page."
    )


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def extract_etf_rows(page) -> list[dict]:
    """
    Parse every <tr role='row'> inside the overview table.
    CCY is read from the AUM <th> header text (e.g. 'AUM ($USD)' → 'USD').
    """
    print("  [step] Detecting currency from AUM column header…")
    raw_header = page.evaluate("""
        () => {
            const container = document.querySelector('#overview') || document.body;
            const ths = Array.from(container.querySelectorAll('table thead th'));
            const aum_th = ths[5] || ths.find(th => th.innerText.includes('AUM'));
            return aum_th ? aum_th.innerText.trim() : '';
        }
    """)
    ccy = parse_ccy_from_header(raw_header) if raw_header else DEFAULT_CCY
    print(f"    → AUM header: {raw_header!r}  →  CCY: {ccy}")

    print("  [step] Extracting ETF rows from DOM…")
    today_str = date.today().isoformat()

    raw_rows: list[dict] = page.evaluate(f"""
        () => {{
            const BASE    = "{BASE_URL}";
            const CCY     = "{ccy}";
            const TODAY   = "{today_str}";
            const resolve = href => (!href ? '' : href.startsWith('http') ? href : BASE + href);

            const container = document.querySelector('#overview') || document.body;
            const rows = Array.from(
                container.querySelectorAll('table.ps_etf_tables__table tbody tr[role="row"]')
            );
            const finalRows = rows.length > 0 ? rows
                : Array.from(container.querySelectorAll('table tbody tr[role="row"]'));

            return finalRows.map(tr => {{
                const tds = Array.from(tr.querySelectorAll('td'));
                const get = i => (tds[i] ? tds[i].innerText.trim() : '');

                const nameAnchor = tds[0] && tds[0].querySelector('a');
                const fund_name  = nameAnchor ? nameAnchor.innerText.trim() : get(0);
                const fund_url   = resolve(nameAnchor ? nameAnchor.getAttribute('href') : null);

                const fsAnchor      = tds[6] && tds[6].querySelector('a');
                const factsheet_url = resolve(fsAnchor ? fsAnchor.getAttribute('href') : null);

                return {{
                    fund_name,
                    fund_url,
                    base_code           : get(1),
                    isin                : get(2),
                    sfdr_classification : get(3),
                    ter_pct             : get(4),
                    aum_usd             : get(5).replace(/,/g, '').trim(),
                    ccy                 : CCY,
                    factsheet_url,
                    scrape_date         : TODAY,
                }};
            }}).filter(r => r.isin.length > 0);
        }}
    """)

    print(f"    → Extracted {len(raw_rows)} ETF rows.")
    return raw_rows


# ─────────────────────────────────────────────────────────────────────────────
# XLSX WRITER
# ─────────────────────────────────────────────────────────────────────────────

def write_xlsx(rows: list[dict], save_dir: Path) -> Path:
    today_str = date.today().isoformat()
    filepath  = save_dir / f"ark_etfs_{today_str}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ARK ETFs"

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper().replace("_", " "))
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for row_idx, record in enumerate(rows, start=2):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else None
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
            if fill:
                cell.fill = fill

    for col_idx in range(1, len(COLUMNS) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(rows) + 2)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"
    wb.save(filepath)
    return filepath


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def _run(headless: bool = True) -> Path:
    save_dir = build_run_output_dir()

    print("=" * 60)
    print("  ARK Invest Europe — ETF Scraper")
    print("=" * 60)
    print(f"  URL      : {URL}")
    print(f"  Save dir : {save_dir}")
    print(f"  Headless : {headless}")
    print("=" * 60)

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=headless,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1440, "height": 900},
            locale="en-GB",
            timezone_id="Europe/London",
        )
        context.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
        page = context.new_page()

        print("\n[1/5] Opening page…")
        page.goto(URL, wait_until="domcontentloaded", timeout=PAGE_LOAD_TIMEOUT)
        page.wait_for_timeout(2_500)

        print("[2/5] Cookie banner…")
        dismiss_cookie_banner(page)

        print("[3/5] Investor gate…")
        dismiss_investor_gate(page)
        page.wait_for_timeout(1_000)

        print("[4/5] Waiting for ETF table…")
        wait_for_fund_table(page)
        page.wait_for_timeout(500)

        print("[5/5] Extracting data…")
        rows = extract_etf_rows(page)

        context.close()
        browser.close()

    if not rows:
        raise RuntimeError(
            "No ETF rows extracted. Run with --headed to debug."
        )

    out_path = write_xlsx(rows, save_dir)

    print()
    print("=" * 60)
    print("  ✅  Scrape complete!")
    print(f"  📊  {len(rows)} ETFs extracted.")
    print(f"  📁  {out_path}")
    print("=" * 60)
    return out_path


async def download_ark_file() -> Path:
    return await asyncio.to_thread(_run, True)


def main() -> None:
    args = parse_args()
    _run(headless=not args.headed)


if __name__ == "__main__":
    main()
