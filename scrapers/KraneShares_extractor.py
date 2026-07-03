"""
Scrape KraneShares UCITS ETF data from kraneshares.eu for a fixed list of ISINs.

Mandatory output columns (one row per ISIN):
    isin | etf_name | issuer | aum_millions | ccy | ter | date

How it works
------------
The 10 target ISINs are Irish-domiciled UCITS share classes.  Multiple ISINs
can belong to the same underlying fund (e.g. KWEB has USD / EUR / GBP hedged
classes).  Each fund has a dedicated page on kraneshares.eu at a known slug.

Strategy:
1.  Use a hardcoded ISIN → (slug, ccy) map so we know exactly which page to
    visit for each share class and what currency it trades in.
2.  Scrape each *unique* fund page once (Playwright, same browser context to
    avoid 403s) and extract: fund name, TER, Net Assets from the Fund Details
    table, plus the primary ISIN shown on the page.
3.  For KWEB share classes (multiple ISINs, one page) we keep the Net Assets
    from the page but assign each ISIN its correct CCY from the map.
4.  AUM is taken from Net Assets on the page and converted to USD millions.
5.  Write one output row per ISIN.
"""

import asyncio
import os
import re
from datetime import datetime
from pathlib import Path

from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ISSUER   = "KraneShares"
BASE_URL = "https://kraneshares.eu/etf"

BASE_DIR   = Path(__file__).resolve().parents[1]
OUTPUT_DIR = BASE_DIR / "providers" / "kraneshares"
TIMEOUT_MS = 60_000
RUN_FOLDER_ENV_VAR = "ETF_PIPELINE_RUN_FOLDER"

# ---------------------------------------------------------------------------
# ISIN → (page_slug, currency, share_class_label)
# Each tuple uniquely identifies how to retrieve and label a share class.
# ---------------------------------------------------------------------------
ISIN_MAP: dict[str, dict] = {
    # KraneShares ICBCUBS S&P China 500 Index UCITS ETF
    "IE0001QF56M0": {"slug": "chinln",  "ccy": "USD", "label": "USD"},

    # KraneShares Global Humanoid Robotics and Physical AI Index UCITS ETF
    # (two different ISINs have appeared for this fund; both map to koidln)
    "IE0009ZB3ZX2": {"slug": "koidln",  "ccy": "USD", "label": "USD"},
    "IE000O6Z73N7": {"slug": "koidln",  "ccy": "USD", "label": "USD"},

    # KraneShares CSI China Internet UCITS ETF – share classes
    "IE00BFXR7892": {"slug": "kwebln",  "ccy": "USD", "label": "USD"},
    "IE00BFXR7900": {"slug": "kwebln",  "ccy": "EUR", "label": "EUR"},
    "IE000K3YPA16": {"slug": "kwebln",  "ccy": "EUR", "label": "EUR Hedged"},
    "IE000CD5SH30": {"slug": "kwebln",  "ccy": "GBP", "label": "GBP Hedged"},
    "IE00BMW13836": {"slug": "kwebln",  "ccy": "EUR", "label": "EUR (Borsa Italiana)"},

    # KraneShares ICBCUBS SSE STAR Market 50 Index UCITS ETF
    "IE00BKPJY434": {"slug": "kstrln",  "ccy": "USD", "label": "USD"},

    # KraneShares Electric Vehicles & Future Mobility Screened UCITS ETF
    "IE000YUAPTQ0": {"slug": "karsln",  "ccy": "USD", "label": "USD"},
}

# Fund Details table label sets (lowercase, stripped)
TER_LABELS    = {"total annual fund operating expense", "gross expense ratio",
                 "expense ratio", "ongoing charges", "ter"}
ASSETS_LABELS = {"net assets", "total net assets", "aum", "fund size"}
NAME_LABELS   = {"fund name", "sub-fund name"}

# Exchange / CCY hint labels (fallback – we trust ISIN_MAP first)
EXCHANGE_CCY = {
    "london stock exchange": "GBP",
    "euronext amsterdam":    "EUR",
    "borsa italiana":        "EUR",
    "deutsche börse":        "EUR",
    "xetra":                 "EUR",
    "default":               "USD",
}


# ---------------------------------------------------------------------------
# Output path helpers
# ---------------------------------------------------------------------------

def build_run_output_dir(base_dir: Path) -> Path:
    run_folder_name = os.environ.get(RUN_FOLDER_ENV_VAR)
    if run_folder_name:
        output_dir = base_dir / run_folder_name
    else:
        run_date = datetime.now().strftime("%Y-%m-%d")
        output_dir = base_dir / run_date
        os.environ[RUN_FOLDER_ENV_VAR] = output_dir.name
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def build_output_path() -> Path:
    return build_run_output_dir(OUTPUT_DIR) / "kraneshares_ucits_export.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _norm(text: str) -> str:
    """Lowercase, strip, remove trailing punctuation."""
    return text.strip().lower().rstrip("*:").strip()


def parse_aum_millions(raw: str) -> str:
    """
    Convert a Net Assets string such as '$5,875,743' or '€125,000,000'
    to AUM in millions (float string, 2 dp).  Returns '' if unparseable.
    Note: values on kraneshares.eu are in USD unless the label says otherwise.
    """
    cleaned = re.sub(r"[^\d.]", "", raw)
    try:
        value = float(cleaned)
        return f"{value / 1_000_000:.2f}"
    except ValueError:
        return ""


# ---------------------------------------------------------------------------
# Core scraper – visits one fund page and extracts the Fund Details table
# ---------------------------------------------------------------------------

async def scrape_fund_page(context, slug: str) -> dict:
    """
    Visit https://kraneshares.eu/etf/<slug>/ and return:
        {name, ter_pct, net_assets_raw, primary_isin}
    Falls back to empty strings on any failure.
    """
    url = f"{BASE_URL}/{slug}/"
    result = {
        "name":           "",
        "ter_pct":        "",
        "net_assets_raw": "",
        "primary_isin":   "",
    }

    page = await context.new_page()
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=TIMEOUT_MS)
        await page.wait_for_timeout(2_500)

        # Collect (label, value) pairs from all <table> rows on the page
        pairs: list[tuple[str, str]] = []

        rows = await page.locator("table tr").all()
        for row in rows:
            cells = await row.locator("td, th").all()
            if len(cells) >= 2:
                label = _norm(await cells[0].inner_text())
                value = (await cells[1].inner_text()).strip()
                pairs.append((label, value))

        # Also check <dt>/<dd> definition-list pairs
        dts = await page.locator("dt").all()
        dds = await page.locator("dd").all()
        for dt, dd in zip(dts, dds):
            pairs.append((_norm(await dt.inner_text()),
                          (await dd.inner_text()).strip()))

        for label, value in pairs:
            if label in TER_LABELS and not result["ter_pct"]:
                result["ter_pct"] = value.replace("%", "").strip().rstrip("*").strip()
            elif label in ASSETS_LABELS and not result["net_assets_raw"]:
                result["net_assets_raw"] = value
            elif label in {"isin", "primary isin"} and not result["primary_isin"]:
                result["primary_isin"] = value.strip()
            elif label in NAME_LABELS and not result["name"]:
                result["name"] = value.strip()

        # Fund name: fall back to the page <h1>
        if not result["name"]:
            h1 = page.locator("h1").first
            if await h1.count():
                result["name"] = (await h1.inner_text()).strip()
                # Remove trailing ticker in brackets e.g. "(KARS)"
                result["name"] = re.sub(r"\s*\([A-Z]{2,6}\)\s*$", "",
                                        result["name"]).strip()

        # ISIN: also look in the page subtitle / meta text (pattern US… or IE…)
        if not result["primary_isin"]:
            content = await page.content()
            m = re.search(r"\b(IE[A-Z0-9]{10})\b", content)
            if m:
                result["primary_isin"] = m.group(1)

    except PlaywrightTimeoutError:
        print(f"    WARNING: timeout on {url}")
    except Exception as exc:
        print(f"    WARNING: error on {url}: {exc}")
    finally:
        await page.close()

    return result


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------

def write_xlsx(records: list[dict], output_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    COLS = [
        ("isin",         "ISIN"),
        ("etf_name",     "ETF Name"),
        ("issuer",       "Issuer"),
        ("aum_millions", "AUM (Millions USD)"),
        ("ccy",          "CCY"),
        ("ter",          "TER (%)"),
        ("date",         "Date"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "KraneShares UCITS ETFs"

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1F3864")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (_, label) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align

    for row_idx, rec in enumerate(records, 2):
        for col_idx, (key, _) in enumerate(COLS, 1):
            value = rec.get(key, "")
            cell  = ws.cell(row=row_idx, column=col_idx)
            if key in ("aum_millions", "ter") and value not in ("", None):
                try:
                    cell.value = float(value)
                    continue
                except ValueError:
                    pass
            cell.value = value

    for col_idx, (key, label) in enumerate(COLS, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(label)
        for rec in records:
            max_len = max(max_len, len(str(rec.get(key, "") or "")))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"
    wb.save(output_path)


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

async def download_kraneshares_ucits() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")

    # Deduplicate slugs so we only scrape each fund page once
    unique_slugs: list[str] = list(dict.fromkeys(
        info["slug"] for info in ISIN_MAP.values()
    ))

    print(f"[1/3] Scraping {len(unique_slugs)} unique UCITS fund pages ...")

    slug_data: dict[str, dict] = {}

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
        )
        context = await browser.new_context(
            locale="en-GB",
            timezone_id="Europe/London",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1440, "height": 1200},
        )

        for slug in unique_slugs:
            print(f"    Fetching /etf/{slug}/ ...")
            slug_data[slug] = await scrape_fund_page(context, slug)

        await browser.close()

    # ── 2. Build one output row per ISIN ─────────────────────────────────
    print("[2/3] Building output rows ...")
    records: list[dict] = []
    missing_name = missing_ter = missing_aum = 0

    for isin, info in ISIN_MAP.items():
        slug   = info["slug"]
        ccy    = info["ccy"]
        label  = info["label"]
        page   = slug_data.get(slug, {})

        base_name = page.get("name", "")
        # Append share-class label to name when there are multiple classes
        # for the same fund (KWEB has USD / EUR / GBP / hedged variants)
        isin_count_for_slug = sum(
            1 for v in ISIN_MAP.values() if v["slug"] == slug
        )
        if isin_count_for_slug > 1 and label:
            etf_name = f"{base_name} ({label})" if base_name else ""
        else:
            etf_name = base_name

        ter       = page.get("ter_pct", "")
        aum_m     = parse_aum_millions(page.get("net_assets_raw", ""))

        if not etf_name: missing_name += 1
        if not ter:       missing_ter  += 1
        if not aum_m:     missing_aum  += 1

        records.append({
            "isin":         isin,
            "etf_name":     etf_name,
            "issuer":       ISSUER,
            "aum_millions": aum_m,
            "ccy":          ccy,
            "ter":          ter,
            "date":         today,
        })

    # ── 3. Write XLSX ─────────────────────────────────────────────────────
    print("[3/3] Writing XLSX ...")
    output_path = build_output_path()
    write_xlsx(records, output_path)

    print(f"\n{'='*60}")
    print(f"  ISINs exported   : {len(records)}")
    print(f"  Missing Name     : {missing_name}")
    print(f"  Missing TER      : {missing_ter}")
    print(f"  Missing AUM      : {missing_aum}")
    print(f"  Date             : {today}")
    print(f"  Output file      : {output_path}")
    print(f"{'='*60}")

    return output_path


if __name__ == "__main__":
    saved = asyncio.run(download_kraneshares_ucits())
    print(f"\nDone! Open your file at: {saved.resolve()}")