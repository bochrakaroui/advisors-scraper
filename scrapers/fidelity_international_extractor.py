"""Download Fidelity International ETF data from Fidelity Luxembourg factsheets.

Workflow:
  1. Read Fidelity International ISINs from ``ISIN-list.xlsx``.
  2. Build a public factsheet URL for each ISIN:
       https://www.fidelity.lu/funds/factsheet/{ISIN}
  3. Fetch the stable ``FundData.json`` payload that the factsheet page uses.
  4. Save one provider-specific raw snapshot JSON.

Output: providers/fidelity/<YYYY-MM-DD>/fidelity_etf_export.json
"""

from __future__ import annotations

import asyncio
import json
import logging
import os
import re
import time
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

import requests
from openpyxl import load_workbook


ISSUER = "Fidelity International"
FACTSHEET_URL_TEMPLATE = "https://www.fidelity.lu/funds/factsheet/{isin}"
FUNDDATA_URL = "https://www.fidelity.lu/api/ce/fdh/FundData.json"
FUNDDATA_PARAMS = {
    "countries": "lu",
    "country": "lu",
    "languages": "en",
    "language": "en",
    "channels": "ce.private-investor",
    "channel": "ce.private-investor",
}
FUNDDATA_FALLBACK_COUNTRIES = ("de", "gb", "ie")

REQUEST_DELAY_S = 0.35

BASE_DIR = Path(__file__).resolve().parents[1]
OUTPUT_DIR = BASE_DIR / "providers" / "fidelity"
ISIN_FILTER_PATH = BASE_DIR / "ISIN-list.xlsx"
RUN_FOLDER_ENV_VAR = "ETF_PIPELINE_RUN_FOLDER"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-GB,en;q=0.9",
    "Accept": "application/json,text/plain,*/*",
}

SESSION = requests.Session()
SESSION.headers.update(HEADERS)

HEADER_NORMALIZE_PATTERN = re.compile(r"[^a-z0-9]+")
INTERNAL_SPACE_PATTERN = re.compile(r"\s+")
INVISIBLE_ISIN_CHARACTERS = ("\u00A0", "\u2007", "\u202F", "\u200B", "\uFEFF")


def build_run_output_dir(base_dir: Path, run_date: str) -> Path:
    run_folder_name = os.environ.get(RUN_FOLDER_ENV_VAR)
    if run_folder_name:
        output_dir = base_dir / run_folder_name
    else:
        output_dir = base_dir / run_date
        os.environ[RUN_FOLDER_ENV_VAR] = output_dir.name

    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def setup_logging() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s", force=True)


def timestamp_now() -> datetime:
    return datetime.now()


def build_output_path(now: datetime) -> Path:
    return build_run_output_dir(OUTPUT_DIR, now.strftime("%Y-%m-%d")) / "fidelity_etf_export.json"


def write_json(path: Path, payload: object) -> None:
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def clean_text(value: object | None) -> str:
    if value is None:
        return ""
    cleaned = str(value).replace("\u00ad", "").strip()
    cleaned = INTERNAL_SPACE_PATTERN.sub(" ", cleaned)
    return "" if cleaned in {"", "-", "--", "- ", " -"} else cleaned


def normalize_isin(value: object | None) -> str:
    if value is None:
        return ""
    normalized = str(value)
    for invisible_character in INVISIBLE_ISIN_CHARACTERS:
        normalized = normalized.replace(invisible_character, "")
    normalized = normalized.strip().upper()
    normalized = INTERNAL_SPACE_PATTERN.sub("", normalized)
    return normalized


def canonicalize_header(value: object | None) -> str:
    return HEADER_NORMALIZE_PATTERN.sub("", clean_text(value).lower())


def format_decimal(value: Decimal, places: int = 2) -> str:
    quantized = value.quantize(Decimal("1." + ("0" * places)), rounding=ROUND_HALF_UP)
    return format(quantized, f".{places}f")


def percentage_to_bps(raw_value: object | None) -> str:
    cleaned = clean_text(raw_value).replace("%", "").replace(",", ".")
    if not cleaned:
        return ""
    try:
        percentage = Decimal(cleaned)
    except InvalidOperation:
        return ""
    return format_decimal(percentage * Decimal("100"), places=2)


def amount_to_millions(raw_value: object | None) -> str:
    cleaned = clean_text(raw_value).replace(",", "")
    if not cleaned:
        return ""
    try:
        amount = Decimal(cleaned)
    except InvalidOperation:
        return ""
    return format_decimal(amount / Decimal("1000000"), places=2)


def load_fidelity_isins(path: Path) -> list[str]:
    if not path.exists():
        raise FileNotFoundError(f"Required ISIN source workbook not found: {path}")

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except PermissionError as exc:
        raise PermissionError(
            f"Cannot read the ISIN filter file because it is locked: {path}. "
            "Please close ISIN-list.xlsx and try again."
        ) from exc

    provider_column_index: int | None = None
    isin_column_index: int | None = None
    ordered_isins: list[str] = []
    seen_isins: set[str] = set()

    worksheet = workbook[workbook.sheetnames[0]]
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if row_index == 1:
            normalized_headers = [canonicalize_header(value) for value in row]
            for index, header in enumerate(normalized_headers):
                if header == "provider":
                    provider_column_index = index
                elif header in {"isin", "isincode"}:
                    isin_column_index = index
            if provider_column_index is None or isin_column_index is None:
                raise ValueError("Could not detect Provider and ISIN columns in ISIN-list.xlsx")
            continue

        provider = clean_text(row[provider_column_index]) if provider_column_index < len(row) else ""
        if provider.lower() != ISSUER.lower():
            continue

        isin = normalize_isin(row[isin_column_index]) if isin_column_index < len(row) else ""
        if not isin or isin in seen_isins:
            continue

        seen_isins.add(isin)
        ordered_isins.append(isin)

    if not ordered_isins:
        raise ValueError("No Fidelity International ISINs were found in ISIN-list.xlsx")

    return ordered_isins


def extract_json_fields(payload: dict[str, object], isin: str) -> dict[str, str]:
    record = payload.get(isin, {})
    if not isinstance(record, dict):
        return {}

    head_fund_facts = record.get("headFundFacts", {})
    share_class_facts = record.get("shareClassFacts", {})
    price_data = record.get("priceData", {})
    performance = record.get("performance", {})
    comparators = performance.get("comparators", {}) if isinstance(performance, dict) else {}
    objective = record.get("objective", {})

    if not isinstance(head_fund_facts, dict):
        head_fund_facts = {}
    if not isinstance(share_class_facts, dict):
        share_class_facts = {}
    if not isinstance(price_data, dict):
        price_data = {}
    if not isinstance(comparators, dict):
        comparators = {}
    comparator_items = comparators.get("items", {}) if isinstance(comparators, dict) else {}
    if not isinstance(comparator_items, dict):
        comparator_items = {}
    if not isinstance(objective, dict):
        objective = {}

    aum_raw = (
        share_class_facts.get("assetsUnderManagement")
        or head_fund_facts.get("assetsUnderManagement")
    )
    ccy = (
        clean_text(share_class_facts.get("currencyName"))
        or clean_text(share_class_facts.get("fundCurrency"))
        or clean_text(price_data.get("currency"))
    )

    return {
        "etf_name": clean_text(record.get("displayName")) or clean_text(share_class_facts.get("fundName")),
        "issuer": ISSUER,
        "isin": normalize_isin(
            share_class_facts.get("isin")
            or record.get("id")
            or isin
        ),
        "ccy": ccy,
        "ter_bps": percentage_to_bps(share_class_facts.get("ongoingChargesFigure")),
        "aum_mn": amount_to_millions(aum_raw),
        "date": clean_text(share_class_facts.get("fundAumEffectiveDate")) or clean_text(price_data.get("date")),
        "fund_type": clean_text(share_class_facts.get("fundType")),
        "share_class_name": clean_text(share_class_facts.get("displayName")) or clean_text(record.get("displayName")),
        "distribution_type": clean_text(share_class_facts.get("distributionType")),
        "share_class": clean_text(share_class_facts.get("shareClass")),
        "primary_ticker": clean_text(share_class_facts.get("primaryTicker")),
        "index_name": clean_text(comparator_items.get("marketIndex")),
        "index_ticker": clean_text(comparator_items.get("indexTicker")),
        "nav": clean_text(price_data.get("nav", {}).get("value") if isinstance(price_data.get("nav"), dict) else ""),
        "nav_date": clean_text(price_data.get("nav", {}).get("date") if isinstance(price_data.get("nav"), dict) else ""),
        "price_currency": clean_text(price_data.get("currency")),
        "objective": clean_text(objective.get("text")),
        "fund_management_approach": clean_text(head_fund_facts.get("fundManagementApproach")),
        "management_company": clean_text(head_fund_facts.get("managementCompany")),
        "fund_domicile": clean_text(head_fund_facts.get("fundDomicile")),
        "product_range": clean_text(share_class_facts.get("productRange")),
    }


def payload_has_record(payload: dict[str, object], isin: str) -> bool:
    record = payload.get(isin)
    return isinstance(record, dict) and bool(record)


def build_funddata_params(isin: str, country_code: str) -> dict[str, str]:
    return {
        **FUNDDATA_PARAMS,
        "countries": country_code,
        "country": country_code,
        "id": isin,
    }


def fetch_funddata_payload(isin: str) -> tuple[dict[str, object], str, str]:
    last_payload: dict[str, object] = {}
    last_url = ""
    last_country = FUNDDATA_PARAMS["country"]
    last_error: Exception | None = None

    country_attempts = (FUNDDATA_PARAMS["country"], *FUNDDATA_FALLBACK_COUNTRIES)
    for country_code in country_attempts:
        try:
            response = SESSION.get(
                FUNDDATA_URL,
                params=build_funddata_params(isin, country_code),
                timeout=45,
            )
            response.raise_for_status()
        except Exception as exc:
            last_error = exc
            continue

        payload = response.json()
        last_payload = payload
        last_url = response.url
        last_country = country_code

        if payload_has_record(payload, isin):
            return payload, last_url, last_country

    if last_error is not None and not last_payload:
        raise last_error
    return last_payload, last_url, last_country


def fetch_isin_row(isin: str) -> dict[str, str]:
    factsheet_url = FACTSHEET_URL_TEMPLATE.format(isin=isin)

    row: dict[str, str] = {
        "etf_name": "",
        "issuer": ISSUER,
        "isin": isin,
        "ccy": "",
        "ter_bps": "",
        "aum_mn": "",
        "date": "",
        "fund_type": "",
        "share_class_name": "",
        "distribution_type": "",
        "share_class": "",
        "primary_ticker": "",
        "index_name": "",
        "index_ticker": "",
        "nav": "",
        "nav_date": "",
        "price_currency": "",
        "objective": "",
        "fund_management_approach": "",
        "management_company": "",
        "fund_domicile": "",
        "product_range": "",
        "factsheet_url": factsheet_url,
        "api_url": "",
        "api_country": "",
        "fetch_status": "pending",
    }

    try:
        payload, api_url, api_country = fetch_funddata_payload(isin)
    except requests.HTTPError as exc:
        status_code = exc.response.status_code if exc.response is not None else "?"
        row["fetch_status"] = f"http_error:{status_code}"
        logging.warning("HTTP %s for Fidelity ISIN %s", status_code, isin)
        return row
    except Exception as exc:
        row["fetch_status"] = f"error:{type(exc).__name__}"
        logging.warning("Could not fetch Fidelity data for %s: %s", isin, exc)
        return row

    extracted = extract_json_fields(payload, isin)
    row.update(extracted)
    row["api_url"] = api_url
    row["api_country"] = api_country
    if row["etf_name"]:
        if api_country != FUNDDATA_PARAMS["country"]:
            logging.info("Fidelity ISIN %s resolved via fallback country context: %s", isin, api_country)
        row["fetch_status"] = "ok"
    elif not payload:
        row["fetch_status"] = "not_found"
    else:
        row["fetch_status"] = "missing_payload"
    return row


def build_snapshot(now: datetime) -> dict[str, object]:
    fidelity_isins = load_fidelity_isins(ISIN_FILTER_PATH)
    logging.info("Loaded %d Fidelity International ISINs from %s", len(fidelity_isins), ISIN_FILTER_PATH)

    listing_rows: list[dict[str, str]] = []
    for index, isin in enumerate(fidelity_isins, start=1):
        logging.info("[%d/%d] Fetching %s", index, len(fidelity_isins), isin)
        listing_rows.append(fetch_isin_row(isin))
        if index < len(fidelity_isins):
            time.sleep(REQUEST_DELAY_S)

    status_counts: dict[str, int] = {}
    for row in listing_rows:
        status = row["fetch_status"]
        status_counts[status] = status_counts.get(status, 0) + 1

    logging.info("Fidelity fetch summary: %s", status_counts)

    return {
        "source": {
            "provider": ISSUER,
            "isin_source": str(ISIN_FILTER_PATH),
            "factsheet_url_template": FACTSHEET_URL_TEMPLATE,
            "funddata_endpoint": FUNDDATA_URL,
        },
        "method": "ISIN workbook filter + Fidelity LU FundData.json per factsheet ISIN",
        "captured_at": now.isoformat(),
        "listing_rows": listing_rows,
    }


def download_snapshot(destination: Path) -> None:
    setup_logging()
    now = timestamp_now()
    snapshot = build_snapshot(now)
    write_json(destination, snapshot)
    logging.info("Data method : %s", snapshot["method"])
    logging.info("Snapshot saved: %s", destination)


async def download_fidelity_file() -> Path:
    now = timestamp_now()
    output_path = build_output_path(now)
    await asyncio.to_thread(download_snapshot, output_path)
    return output_path


def parse_snapshot_rows(path: Path) -> list[dict[str, str]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    return payload.get("listing_rows", [])


def main() -> None:
    output_path = build_output_path(timestamp_now())
    download_snapshot(output_path)


if __name__ == "__main__":
    main()
