"""Scrape Robeco ETF data and save to an XLSX workbook.

Selectors are derived from live DOM inspection of:
  - Listing page : https://www.robeco.com/en-uk/products/etf
  - Detail page  : https://www.robeco.com/en-uk/products/funds/isin-*/robeco-*

Page structure confirmed:
  LISTING
  -------
  • Fund-level  rows  → <h5> inside <li>  (no href – expands share-class rows below it)
  • Share-class rows  → <h6><a href="/en-uk/products/funds/isin-XXXX/...">NAME</a>
                        followed by text "ISIN: XXXXXXXXXXXX"
  • Performance cols  → bare text nodes: 1m%, 3m%, YTD%, 1y%, 5y%
  • Morningstar score → e.g. "9.15 ( 24-06 )"
  • "Fund view" / "Share class view" toggle changes the DOM layout

  DETAIL PAGE  (tab=overview is the default / SSR-rendered tab)
  -----------
  • h1            → full fund name  (e.g. "Robeco 3D Global Equity UCITS ETF USD Acc")
  • h3 (first)    → description / objective
  • "General Facts" <table>  → key/value pairs for ISIN, Bloomberg ticker,
                               product structure, asset class, SFDR, currencies,
                               AUM, share-class size, inception date, ongoing charges,
                               use of income, management company …
  • "Trading Information" section  → lists exchanges  (LSE, SIX, XETRA, …)
  • "Performance" <table>         → 1m, 3m, YTD, 1y, since-inception vs index
  • "Costs" <table>               → management fee, transaction costs
  • "Class and codes" aside       → ISIN, Bloomberg, Index (benchmark)
  • SFDR label                    → plain text "Article 8" after tooltip trigger
"""

import asyncio
import os
import re
from datetime import datetime
from pathlib import Path

from playwright.async_api import TimeoutError as PlaywrightTimeoutError, async_playwright


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
URL = "https://www.robeco.com/en-uk/products/etf"
BASE_DIR = Path(__file__).resolve().parents[1]
OUTPUT_DIR = BASE_DIR / "providers" / "Robeco"
TIMEOUT_MS = 120_000
RUN_FOLDER_ENV_VAR = "ETF_PIPELINE_RUN_FOLDER"


# ---------------------------------------------------------------------------
# Output path
# ---------------------------------------------------------------------------
def build_run_output_dir(base_dir: Path) -> Path:
    run_folder_name = os.environ.get(RUN_FOLDER_ENV_VAR)
    if run_folder_name:
        output_dir = base_dir / run_folder_name
    else:
        run_date = datetime.now().strftime("%Y-%m-%d")
        output_dir = base_dir / run_date
        os.environ[RUN_FOLDER_ENV_VAR] = output_dir.name
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def build_output_path() -> Path:
    return build_run_output_dir(OUTPUT_DIR) / "robeco_etf_export.xlsx"


# ---------------------------------------------------------------------------
# Overlay / cookie / disclaimer dismissal
# ---------------------------------------------------------------------------
async def dismiss_overlays(page) -> None:
    """
    Robeco UK shows:
      1. A cookie-consent bar   – OneTrust  (#onetrust-accept-btn-handler)
      2. A professional-investor disclaimer modal on the ETF selector page.
         The modal body contains text about MiFID professional investors.
         Confirmed buttons from DOM: text varies; try multiple selectors.
    """
    # 1. Cookie banner (OneTrust is the most common CMP on robeco.com)
    for sel in [
        "#onetrust-accept-btn-handler",
        "button.onetrust-accept-btn-handler",
        "button:has-text('Accept All Cookies')",
        "button:has-text('Accept all')",
    ]:
        try:
            btn = page.locator(sel).first
            if await btn.is_visible(timeout=4_000):
                print(f"    [cookie] clicking: {sel}")
                await btn.click(timeout=6_000)
                await page.wait_for_timeout(1_200)
                break
        except Exception:
            continue

    # 2. Professional-investor disclaimer
    #    The modal is rendered as a full-page overlay with a prominent button.
    #    From the page HTML the text around the button reads
    #    "NOT FOR RETAIL CLIENTS … professional investors …"
    #    Button candidates (in order of specificity):
    for sel in [
        # Robeco-specific data attributes (check DevTools → might be present)
        "button[data-profile='PROFESSIONAL']",
        "button[data-profile='INSTIT']",
        # Generic text matches (SSR-rendered text confirmed from fetch)
        "button:has-text('Professional investor')",
        "button:has-text('I am a professional investor')",
        "button:has-text('I confirm')",
        "button:has-text('Confirm')",
        "button:has-text('I Agree')",
        "button:has-text('Agree and continue')",
        # Fallback: any visible button inside a modal/overlay container
        ".modal button, [class*='disclaimer'] button, [class*='overlay'] button",
    ]:
        try:
            btn = page.locator(sel).first
            if await btn.is_visible(timeout=3_000):
                print(f"    [disclaimer] clicking: {sel}")
                await btn.click(timeout=6_000)
                await page.wait_for_timeout(2_000)
                break
        except Exception:
            continue

    # JS safety net: strip any residual backdrop / scroll-lock
    await page.evaluate("""
        () => {
            document.querySelectorAll('.modal-backdrop, [class*="overlay"], [id*="cookie"]')
                    .forEach(el => el.remove());
            document.body.classList.remove('modal-open');
            document.body.style.overflow = 'auto';
        }
    """)
    await page.wait_for_timeout(400)


# ---------------------------------------------------------------------------
# Listing page  –  collect share-class links + per-row data
# ---------------------------------------------------------------------------

# JS injected into the listing page to read all visible ETF rows.
# Confirmed DOM structure:
#   <li>                                  ← fund group (h5 row)
#     <h5>3D Global Equity UCITS ETF</h5>
#     <span>3 Share classes</span>
#     …performance + price columns…
#     <ul>                                ← share-class sub-list
#       <li>                              ← share-class row (h6 row)
#         <h6><a href="/en-uk/products/funds/isin-IE000Q8N7WY1/…">NAME</a></h6>
#         <span>ISIN: IE000Q8N7WY1</span>
#         …performance columns…
#         <span>6.79</span>              ← morningstar / price score
#         <span>( 24-06 )</span>         ← date
#       </li>
#     </ul>
#   </li>
#
# Performance columns order (confirmed from column headers):
#   1 month | 3 months | YTD | 1 year | 5 years | morningstar | share class | currency | current price* (date)
#
# When "Fund view" is active (default), the share-class rows are NESTED inside
# the fund row.  When "Share class view" is active, they are flat.
# We switch to "Share class view" (adds ?lt=list to URL) for a flat list,
# which is easier to parse.

LISTING_JS = """
() => {
    const results = [];
    const seen = new Set();

    // Every share-class row has an anchor whose href matches /products/funds/isin-
    const anchors = document.querySelectorAll("a[href*='/products/funds/isin-']");
    for (const a of anchors) {
        const href = a.href;
        if (!href || seen.has(href)) continue;
        seen.add(href);

        // ISIN from URL  (pattern: isin-IE000Q8N7WY1)
        const isinM = href.match(/\/isin-([A-Z0-9]{12})/i);
        const isin = isinM ? isinM[1].toUpperCase() : "";

        // Share-class name from the anchor text
        const name = a.innerText.trim();

        // Walk up to the share-class <li> to read sibling data
        const li = a.closest("li");
        const liText = li ? li.innerText : "";

        // ISIN also appears as plain text "ISIN: XXXX" in the same row
        const isinTextM = liText.match(/ISIN[:\\s]+([A-Z0-9]{12})/i);
        const isinFromText = isinTextM ? isinTextM[1].toUpperCase() : "";

        // Performance columns – appear as "X.XX%" strings in the row
        const pctMatches = [...liText.matchAll(/([-\\d]+\\.\\d+)%/g)].map(m => m[1] + "%");

        // Morningstar score / current price – a float not followed by %
        const priceM = liText.match(/(\\d+\\.\\d{2})(?!%)/);
        const price = priceM ? priceM[1] : "";

        // Price date  –  pattern "( DD-MM )"
        const dateM = liText.match(/\\(\\s*(\\d{2}-\\d{2})\\s*\\)/);
        const priceDate = dateM ? dateM[1] : "";

        // Fund-group name  –  the <h5> parent of this share-class cluster
        const fundLi = li ? li.closest("ul")?.closest("li") : null;
        const h5 = fundLi ? fundLi.querySelector("h5") : null;
        const fundName = h5 ? h5.innerText.trim() : "";

        // Fund-level description  –  text node after h5 in the parent li
        const descEl = fundLi ? fundLi.querySelector("h5 + *, p, [class*='description']") : null;
        const fundDescription = descEl ? descEl.innerText.trim() : "";

        results.push({
            share_class_name: name,
            fund_name: fundName,
            fund_description: fundDescription,
            isin: isin || isinFromText,
            perf_1m:  pctMatches[0] || "",
            perf_3m:  pctMatches[1] || "",
            perf_ytd: pctMatches[2] || "",
            perf_1y:  pctMatches[3] || "",
            perf_5y:  pctMatches[4] || "",
            price: price,
            price_date: priceDate,
            url: href,
        });
    }
    return results;
}
"""


async def collect_listing_rows(page) -> list[dict]:
    """Switch to flat share-class view then extract all rows."""

    # Switch to share-class (flat) view  →  appends ?lt=list
    flat_url = URL + "?lt=list"
    print(f"    Navigating to flat share-class view: {flat_url}")
    await page.goto(flat_url, wait_until="domcontentloaded", timeout=TIMEOUT_MS)
    await page.wait_for_timeout(4_000)
    await dismiss_overlays(page)

    # Wait for share-class anchors to appear
    print("    Waiting for share-class anchors ...")
    try:
        await page.wait_for_selector(
            "a[href*='/products/funds/isin-']",
            timeout=60_000,
            state="attached",
        )
    except PlaywrightTimeoutError:
        print("    WARNING: Timed out waiting for share-class anchors.")

    # Scroll to bottom to trigger lazy-loading
    print("    Scrolling to load all rows ...")
    prev_h = -1
    for _ in range(25):
        h = await page.evaluate("() => document.body.scrollHeight")
        if h == prev_h:
            break
        prev_h = h
        await page.evaluate("() => window.scrollTo(0, document.body.scrollHeight)")
        await page.wait_for_timeout(1_200)

    rows: list[dict] = await page.evaluate(LISTING_JS)
    print(f"    Found {len(rows)} share-class rows on listing page.")
    return rows


# ---------------------------------------------------------------------------
# Detail page scraping
# ---------------------------------------------------------------------------

# JS to read the detail page.
# Confirmed structure (from live fetch of the 3D Global Equity USD Acc page):
#
#   h1                        → full share-class name
#   h3 (first after h1)       → objective / description
#   "Class and codes" section → ISIN, Bloomberg, Index  (in <dt>/<dd> or similar)
#   SFDR label                → "Article 8" text near ".sfdr" or after tooltip
#   "General Facts" table     → <table> with <tr><td>label</td><td>value</td></tr>
#     Confirmed keys (from fetched HTML):
#       Primary Ticker, ISIN, Product Structure, Fund management approach,
#       Investment strategy type, Asset class, SFDR classification,
#       Fund base currency, Share class currency, Total size of fund,
#       Size of shareclass, Shareclass outstanding shares,
#       Share class inception date, Close financial year,
#       Share class ongoing charges, Use of income, Management company
#   "Trading Information" section → plain text lines per exchange
#   "Performance" table   → 1m, 3m, YTD, 1y, since-inception  (fund + index)
#   "Costs" table         → management fee, transaction costs

DETAIL_JS = """
() => {
    // ── helpers ──────────────────────────────────────────────────────────
    const txt = (sel, root) => {
        const el = (root || document).querySelector(sel);
        return el ? el.innerText.trim() : "";
    };
    const allTxt = (sel, root) =>
        [...(root || document).querySelectorAll(sel)]
            .map(e => e.innerText.trim()).filter(Boolean);

    // ── name & description ───────────────────────────────────────────────
    const name = txt("h1");
    // First <h3> that follows the <h1> is the fund objective
    const h3s = [...document.querySelectorAll("h3")];
    const objective = h3s.find(h => h.innerText.trim().length > 20)?.innerText.trim() || "";

    // ── ISIN from page title / URL ───────────────────────────────────────
    const isinM = document.title.match(/([A-Z]{2}[A-Z0-9]{10})/);
    const isin = isinM ? isinM[1] : "";

    // ── "Class and codes" sidebar ────────────────────────────────────────
    // Confirmed labels: Asset class, ISIN, Bloomberg
    // Also nearby: Index label (e.g. "MSCI World Index")
    const codeBlock = {};
    // Try dt/dd pairs
    for (const dt of document.querySelectorAll("dt")) {
        const dd = dt.nextElementSibling;
        if (dd && dd.tagName === "DD") {
            codeBlock[dt.innerText.trim()] = dd.innerText.trim();
        }
    }
    // Try generic label/value pairs in the aside / sidebar
    for (const el of document.querySelectorAll("[class*='codes'] li, [class*='class-code'] li")) {
        const label = el.querySelector("span:first-child, strong, b")?.innerText.trim() || "";
        const value = el.querySelector("span:last-child, em")?.innerText.trim() || "";
        if (label && value) codeBlock[label] = value;
    }

    const bloomberg = codeBlock["Bloomberg"] || "";
    const index     = codeBlock["Index"] || codeBlock["Benchmark"] || "";

    // ── SFDR article ─────────────────────────────────────────────────────
    // Appears as plain text "Article 8" after a tooltip trigger button
    const sfdrM = document.body.innerText.match(/Article\\s+(6|8|9)/);
    const sfdr  = sfdrM ? "Article " + sfdrM[1] : "";

    // ── "General Facts" table ─────────────────────────────────────────────
    // All tables on the page; find the one whose first row label is "Primary Ticker"
    // or "ISIN" or "Asset class" (first column = label, second = value)
    const facts = {};
    for (const table of document.querySelectorAll("table")) {
        const rows = table.querySelectorAll("tr");
        let isFactsTable = false;
        for (const row of rows) {
            const cells = row.querySelectorAll("td, th");
            if (cells.length >= 2) {
                const label = cells[0].innerText.trim();
                const value = cells[1].innerText.trim();
                // Signature fields of the General Facts table
                if (["Primary Ticker","ISIN","Asset class","SFDR classification",
                     "Fund base currency","Share class inception date",
                     "Share class ongoing charges","Management company"].includes(label)) {
                    isFactsTable = true;
                }
                if (label) facts[label] = value;
            }
        }
        // Stop after we've found and parsed the General Facts table
        if (isFactsTable) break;
    }

    // ── Performance table ─────────────────────────────────────────────────
    // Header row: "Per period | Fund | Index"
    // Data rows : "1 month | -6.46% | -6.37%"  etc.
    const perf = {};
    for (const table of document.querySelectorAll("table")) {
        const headers = [...table.querySelectorAll("th")].map(th => th.innerText.trim());
        if (headers.join(" ").toLowerCase().includes("fund") &&
            headers.join(" ").toLowerCase().includes("index")) {
            for (const row of table.querySelectorAll("tr")) {
                const cells = [...row.querySelectorAll("td")];
                if (cells.length >= 2) {
                    const period = cells[0].innerText.trim();
                    const fund   = cells[1]?.innerText.trim() || "";
                    const idx    = cells[2]?.innerText.trim() || "";
                    if (period) {
                        perf["perf_fund_" + period]  = fund;
                        perf["perf_index_" + period] = idx;
                    }
                }
            }
        }
    }

    // ── Costs table ──────────────────────────────────────────────────────
    // Header: "Cost of this fund | Percentage"
    const costs = {};
    for (const table of document.querySelectorAll("table")) {
        const headers = [...table.querySelectorAll("th")].map(th => th.innerText.trim().toLowerCase());
        if (headers.some(h => h.includes("cost")) || headers.some(h => h.includes("percentage"))) {
            for (const row of table.querySelectorAll("tr")) {
                const cells = [...row.querySelectorAll("td")];
                if (cells.length >= 2) {
                    const label = cells[0].innerText.replace(/\\s+/g," ").trim().slice(0,60);
                    const value = cells[1].innerText.trim();
                    if (label) costs[label] = value;
                }
            }
        }
    }

    // ── Trading information (exchanges) ───────────────────────────────────
    // Appears as a section with plain text lines: "LSE - United Kingdom", etc.
    const tradingSection = (() => {
        for (const h of document.querySelectorAll("h3, h4")) {
            if (/trading information/i.test(h.innerText)) {
                let el = h.nextElementSibling;
                const parts = [];
                while (el && !["H2","H3","H4"].includes(el.tagName)) {
                    const t = el.innerText.trim();
                    if (t) parts.push(t);
                    el = el.nextElementSibling;
                }
                return parts.join(" | ");
            }
        }
        return "";
    })();

    // ── Fund managers ─────────────────────────────────────────────────────
    const managers = allTxt("[class*='manager'] h3, [class*='manager'] h4, [class*='manager'] strong").join(", ");

    // ── Assemble output ───────────────────────────────────────────────────
    return {
        // Core identity
        name:                      name,
        isin:                      isin || facts["ISIN"] || "",
        bloomberg_ticker:          bloomberg || facts["Primary Ticker"] || "",
        description:               objective,

        // Classification
        asset_class:               facts["Asset class"] || "",
        sfdr_article:              sfdr,
        product_structure:         facts["Product Structure"] || "",
        management_approach:       facts["Fund management approach"] || "",
        strategy_type:             facts["Investment strategy type"] || "",

        // Currencies
        fund_base_currency:        facts["Fund base currency"] || "",
        share_class_currency:      facts["Share class currency"] || "",

        // Size
        total_fund_size:           facts["Total size of fund"] || "",
        share_class_size:          facts["Size of shareclass"] || "",
        outstanding_shares:        facts["Shareclass outstanding shares"] || "",

        // Dates & income
        inception_date:            facts["Share class inception date"] || "",
        close_financial_year:      facts["Close financial year"] || "",
        use_of_income:             facts["Use of income"] || "",

        // Fees
        ongoing_charges:           facts["Share class ongoing charges"] || costs["Management fee"] || "",
        transaction_costs:         costs["Transaction costs"] || "",
        management_fee:            costs["Management fee"] || "",

        // Benchmark
        benchmark_index:           index || facts["Benchmark"] || "",

        // Operations
        management_company:        facts["Management company"] || "",
        exchanges:                 tradingSection,
        fund_managers:             managers,

        // Performance (from detail page performance table)
        ...perf,

        // URL
        url: window.location.href,

        // Raw dump for debugging
        _raw_facts: JSON.stringify(facts),
        _raw_costs: JSON.stringify(costs),
    };
}
"""


async def scrape_detail(page, url: str) -> dict:
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=TIMEOUT_MS)
        await page.wait_for_timeout(3_500)
        # Wait for the General Facts table to appear
        try:
            await page.wait_for_selector("table", timeout=15_000, state="attached")
        except PlaywrightTimeoutError:
            pass
        return await page.evaluate(DETAIL_JS)
    except Exception as exc:
        print(f"    ERROR scraping {url}: {exc}")
        return {"url": url, "error": str(exc)}


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

COLUMNS = [
    # ── From listing page ─────────────────────────────
    ("Fund Name",               "fund_name"),
    ("Share Class Name",        "share_class_name"),
    ("ISIN",                    "isin"),
    ("Fund Description",        "fund_description"),
    # ── From detail page ──────────────────────────────
    ("Full Name (detail)",      "name"),
    ("Objective",               "description"),
    ("Bloomberg Ticker",        "bloomberg_ticker"),
    ("Asset Class",             "asset_class"),
    ("SFDR Article",            "sfdr_article"),
    ("Product Structure",       "product_structure"),
    ("Management Approach",     "management_approach"),
    ("Strategy Type",           "strategy_type"),
    ("Fund Base Currency",      "fund_base_currency"),
    ("Share Class Currency",    "share_class_currency"),
    ("Total Fund Size",         "total_fund_size"),
    ("Share Class Size",        "share_class_size"),
    ("Outstanding Shares",      "outstanding_shares"),
    ("Inception Date",          "inception_date"),
    ("Close Financial Year",    "close_financial_year"),
    ("Use of Income",           "use_of_income"),
    ("Ongoing Charges",         "ongoing_charges"),
    ("Management Fee",          "management_fee"),
    ("Transaction Costs",       "transaction_costs"),
    ("Benchmark / Index",       "benchmark_index"),
    ("Management Company",      "management_company"),
    ("Exchanges",               "exchanges"),
    ("Fund Managers",           "fund_managers"),
    # ── Performance from listing page ─────────────────
    ("Perf 1M (listing)",       "perf_1m"),
    ("Perf 3M (listing)",       "perf_3m"),
    ("Perf YTD (listing)",      "perf_ytd"),
    ("Perf 1Y (listing)",       "perf_1y"),
    ("Perf 5Y (listing)",       "perf_5y"),
    ("Price (listing)",         "price"),
    ("Price Date (listing)",    "price_date"),
    # ── Performance from detail page ──────────────────
    ("Perf Fund 1 month",       "perf_fund_1 month"),
    ("Perf Index 1 month",      "perf_index_1 month"),
    ("Perf Fund 3 months",      "perf_fund_3 months"),
    ("Perf Index 3 months",     "perf_index_3 months"),
    ("Perf Fund YTD",           "perf_fund_YTD"),
    ("Perf Index YTD",          "perf_index_YTD"),
    ("Perf Fund 1 year",        "perf_fund_1 year"),
    ("Perf Index 1 year",       "perf_index_1 year"),
    ("Perf Fund Since Inception","perf_fund_Since inception"),
    ("Perf Index Since Inception","perf_index_Since inception"),
    # ── Meta ──────────────────────────────────────────
    ("URL",                     "url"),
    ("Raw Facts JSON",          "_raw_facts"),
    ("Error",                   "error"),
]


def save_to_xlsx(records: list[dict], output_path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is required.  pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Robeco ETFs"

    header_fill = PatternFill("solid", fgColor="003781")   # Robeco dark blue
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for col_idx, (header, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    for row_idx, rec in enumerate(records, 2):
        for col_idx, (_, field) in enumerate(COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=rec.get(field, "") or "")

    # Auto-width: cap at 60
    for col_idx in range(1, len(COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, len(records) + 2)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"    Saved {len(records)} rows → {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
async def scrape_robeco_etfs() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
        )
        ctx = await browser.new_context(
            locale="en-GB",
            timezone_id="Europe/London",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1440, "height": 1200},
        )
        page = await ctx.new_page()

        # ── Step 1: listing page ─────────────────────────────────────────
        print("[1/4] Loading Robeco ETF selector ...")
        await page.goto(URL, wait_until="domcontentloaded", timeout=TIMEOUT_MS)
        await page.wait_for_timeout(4_000)

        # ── Step 2: dismiss overlays ─────────────────────────────────────
        print("[2/4] Dismissing overlays ...")
        await dismiss_overlays(page)

        # ── Step 3: collect listing rows ─────────────────────────────────
        print("[3/4] Collecting share-class rows from listing ...")
        listing_rows = await collect_listing_rows(page)

        if not listing_rows:
            print("    WARNING: no rows found – check overlay dismissal.")

        # ── Step 4: scrape each detail page ─────────────────────────────
        print(f"[4/4] Scraping {len(listing_rows)} detail pages ...")
        records: list[dict] = []
        for idx, row in enumerate(listing_rows, 1):
            url = row.get("url", "")
            label = row.get("share_class_name") or row.get("isin") or url[:60]
            print(f"    [{idx}/{len(listing_rows)}] {label}")
            if url:
                detail = await scrape_detail(page, url)
                merged = {**row, **detail}          # detail wins on conflicts
                # Preserve listing-page ISIN if detail missed it
                if not merged.get("isin") and row.get("isin"):
                    merged["isin"] = row["isin"]
                records.append(merged)
            else:
                records.append(row)

        await browser.close()

    final_path = build_output_path()
    print(f"\nSaving to {final_path} ...")
    save_to_xlsx(records, final_path)
    return final_path


async def download_robeco_file() -> Path:
    return await scrape_robeco_etfs()


if __name__ == "__main__":
    saved = asyncio.run(scrape_robeco_etfs())
    print(f"\nDone!  Open: {saved.resolve()}")
